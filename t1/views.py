from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from .models import *
from . import forms
import pandas as pd
import openpyxl
from openpyxl import Workbook
from pandas import ExcelWriter
from django.contrib.auth.hashers import make_password, check_password
from openpyxl.writer.excel import save_virtual_workbook


# Create your views here.




def login(request):
    if(request.method=='POST'):

        userna = request.POST.get('username')
        username = userna.upper()

        if(User.objects.filter(username=username).exists()):
            user = User.objects.get(username=username)
            first_name = user.first_name
            email = user.email
            phone = user.phone
            # random_otp = r''.join(random.choice('0123456789') for i in range(4))
            # phone_otp(random_otp,phone)

            random_otp = '1234'

            hashed_pwd = make_password(random_otp)
            User.objects.filter(username=username).update(password=hashed_pwd)

            return HttpResponseRedirect("/login/user=" + username)

    return render(request,'login.html')

def index(request):

    # form1 = forms.form_test_year()



    data0 = User.objects.get(username=request.user)
    print(data0.department)
    print(data0.first_name)


    data1 = Field.objects.all()
    data2 = Year.objects.all()
    data3 = zip(data1,data2)




    context = {
            'data0':data0,
            'data1':data1,
            'data2':data2,
            'data3':data3


        }

    return render(request,'index.html',context = context)


def faculty_index(request):

    data0 = User.objects.get(username=request.user)
    data1 = Field.objects.all()
    data2 = Year.objects.all()


    context = {
            'data0':data0,
            'data1':data1,
            'data2':data2



        }

    return render(request,'faculty_index.html',context = context)


def principal_second(request,dept):

    data0 = User.objects.get(username=request.user)
    data1 = Field.objects.all()
    data2 = Year.objects.all()

    context = {
            'data0':data0,
            'data1':data1,
            'data2':data2,
            'data3':dept



        }

    return render(request,'faculty_index.html',context = context)



def first_year(request):
    data0 = Year.objects.all()
    data1 = User.objects.get(username=request.user)

    context = {
            'data0':data0,
            'data1':data1
        }

    return render(request,'first_year.html',context=context)




def first_new(request,year):

    if fdp.objects.filter(info=request.user).filter(year=year):
        return HttpResponseRedirect('/first_year')

    else:
  
        form_fdp = forms.form_fdp()
        form_refreshers_course = forms.form_refreshers_course()
        form_sttp = forms.form_sttp()
        form_workshop = forms.form_workshop()

        form_book = forms.form_book()
        form_chapter = forms.form_chapter()

        form_seminar = forms.form_seminar()
        form_technical_talk = forms.form_technical_talk()
        form_session_chair = forms.form_session_chair()

        form_honours = forms.form_honours()

        form_online_courses = forms.form_online_courses()

        form_consultancy = forms.form_consultancy()

        form_guidance = forms.form_phd_guide()
        form_pursuing_details = forms.form_phd_self()

        form_conference = forms.form_conference()
        form_journal = forms.form_journal()

        form_funded_project = forms.form_funding()



        data0 = User.objects.get(username=request.user)
        print(data0.department)
        print(data0.first_name)


        # data1 = Field.objects.all()
        data2 = Year.objects.all()
        # data3 = zip(data1,data2)
 
        if request.method == 'POST':

                
            form_fdp = forms.form_fdp(request.POST)
            form_refreshers_course = forms.form_refreshers_course(request.POST)
            form_sttp = forms.form_sttp(request.POST)
            form_workshop = forms.form_workshop(request.POST)

            form_book = forms.form_book(request.POST)
            form_chapter = forms.form_chapter(request.POST)

            form_seminar = forms.form_seminar(request.POST)
            form_technical_talk = forms.form_technical_talk(request.POST)
            form_session_chair = forms.form_session_chair(request.POST)

            form_honours = forms.form_honours(request.POST)

            form_online_courses = forms.form_online_courses(request.POST)

            form_consultancy = forms.form_consultancy(request.POST)

            form_guidance = forms.form_phd_guide(request.POST)
            form_pursuing_details = forms.form_phd_self(request.POST)

            form_conference = forms.form_conference(request.POST)
            form_journal = forms.form_journal(request.POST)

            form_funded_project = forms.form_funding(request.POST)


            if form_fdp.is_valid():
                

                obj1 = form_fdp.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()

            if form_refreshers_course.is_valid():
                

                obj1 = form_refreshers_course.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()    


            if form_sttp.is_valid():
                

                obj1 = form_sttp.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_workshop.is_valid():
                

                obj1 = form_workshop.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_book.is_valid():
                

                obj1 = form_book.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  



            if form_chapter.is_valid():
                

                obj1 = form_chapter.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  



            if form_seminar.is_valid():
                

                obj1 = form_seminar.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_technical_talk.is_valid():
                

                obj1 = form_technical_talk.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_session_chair.is_valid():
                

                obj1 = form_session_chair.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  



            if form_honours.is_valid():
                

                obj1 = form_honours.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_online_courses.is_valid():
                

                obj1 = form_online_courses.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save() 


            if form_consultancy.is_valid():
                

                obj1 = form_consultancy.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save() 


            if form_guidance.is_valid():
                

                obj1 = form_guidance.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save() 


            if form_pursuing_details.is_valid():
                

                obj1 = form_pursuing_details.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  

            if form_conference.is_valid():
                

                obj1 = form_conference.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_journal.is_valid():
                

                obj1 = form_journal.save(commit=False)
                obj1.info = request.user
                obj1.year = year

                obj1.save()  


            if form_funded_project.is_valid():
                obj4 = form_funded_project.save(commit=False)
                obj4.info = request.user
                obj4.year = year

                obj4.save()

                return HttpResponseRedirect('/first_year')



        context = {
                'data0':data0,
                'data2':data2,
                'current_year':year,
                'form_fdp':form_fdp,
                'form_refreshers_course':form_refreshers_course,
                'form_sttp':form_sttp,
                'form_workshop':form_workshop,
                'form_book':form_book,
                'form_chapter':form_chapter,

                'form_seminar':form_seminar,
                'form_technical_talk':form_technical_talk,
                'form_session_chair':form_session_chair,

                'form_honours':form_honours,
                'form_online_courses':form_online_courses,
                'form_consultancy':form_consultancy,

                'form_guidance':form_guidance,
                'form_pursuing_details':form_pursuing_details,
                'form_conference':form_conference,
                'form_journal':form_journal,

                'form_funded_project':form_funded_project
               
                
                
            }

        return render(request,'first_new.html',context = context)







def principal_first(request):


    data1 = Department.objects.all()

    context = {

        'data1':data1
    }

    return render(request,'principal_first.html',context=context)






def decide(request):

    if request.user.is_assistant_professor():

        return HttpResponseRedirect("/first_year/")

    elif request.user.is_associate_professor():
        return HttpResponseRedirect("/index/")


    elif request.user.is_professor():
        return HttpResponseRedirect("/index/")


    elif request.user.is_hod():

        return HttpResponseRedirect("/index/")


    elif request.user.is_principal():
        return HttpResponseRedirect("/first/")



    elif request.user.is_ao():
        return HttpResponseRedirect("/first/")





def fdp1(request,year):
    form = forms.form_fdp()


    if fdp.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_fdp/'+year)


    if request.method == 'POST':
        # status = User.objects.get(username=request.user)
        form = forms.form_fdp(request.POST)


        if form.is_valid():
            print(year)

            obj = form.save(commit=False)
            obj.info = request.user
            obj.year = year

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'fdp.html',{'form':form})


def fdp1_preview(request,year):


    data1 = fdp.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_fdp.html',context=context)


def fdp1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(fdp.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="fdp_report.xlsx"'
    return response





def ref_course1(request,year):
    form = forms.form_refreshers_course()
    if refreshers_course.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_ref_course/'+year)

    if request.method == 'POST':

        form = forms.form_refreshers_course(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user
            obj.year = year

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'ref_course.html',{'form':form})



def ref_course1_preview(request,year):


    data1 = refreshers_course.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_ref_course.html',context=context)


def ref_course1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(refreshers_course.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ref_course_report.xlsx"'
    return response


def sttp1(request,year):
    form = forms.form_sttp()
    if sttp.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_sttp/'+year)

    if request.method == 'POST':


        form = forms.form_sttp(request.POST)


        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')


    return render(request,'sttp.html',{'form':form})


def sttp1_preview(request,year):


    data1 = sttp.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_sttp.html',context=context)

def sttp1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(sttp.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="sttp_report.xlsx"'
    return response


def book1(request,year):
    form = forms.form_book()
    if book.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_book/'+year)
    if request.method == 'POST':

        form = forms.form_book(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'book.html',{'form':form})

def book1_preview(request,year):


    data1 = book.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_book.html',context=context)

def book1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(book.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="book_report.xlsx"'
    return response


def interaction1(request,year):
    form = forms.form_interaction()
    if interaction.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_interaction/'+year)

    if request.method == 'POST':

        form = forms.form_interaction(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'interaction.html',{'form':form})


def interaction1_preview(request,year):


    data1 = interaction.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_interaction.html',context=context)

def interaction1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(interaction.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="interaction_report.xlsx"'
    return response

def honours1(request,year):
    form = forms.form_honours()
    if honours.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_honours/'+year)
    if request.method == 'POST':

        form = forms.form_honours(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')



    return render(request,'honours.html',{'form':form})





def honours1_preview(request,year):


    data1 = honours.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_honours.html',context=context)


def honours1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(honours.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="honours_report.xlsx"'
    return response



def online_courses1(request,year):
    form = forms.form_online_courses()
    if online_courses.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_online_courses/'+year)
    if request.method == 'POST':

        form = forms.form_online_courses(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'online_courses.html',{'form':form})

def online_courses1_preview(request,year):
    data1 = online_courses.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_online_courses.html',context=context)


def online_courses1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(online_courses.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="online_courses_report.xlsx"'
    return response



def consultancy1(request,year):
    form = forms.form_consultancy()
    if consultancy.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_consultancy/'+year)
    if request.method == 'POST':

        form = forms.form_consultancy(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')



    return render(request,'consultancy.html',{'form':form})

def consultancy1_preview(request,year):
    data1 = consultancy.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_consultancy.html',context=context)

def consultancy1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(consultancy.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="consultancy_report.xlsx"'
    return response


def phd_guide1(request,year):
    form = forms.form_phd_guide()
    if phd_guide.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_phd_guide/'+year)
    if request.method == 'POST':

        form = forms.form_phd_guide(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'phd_guide.html',{'form':form})

def phd_guide1_preview(request,year):
    data1 = phd_guide.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_phd_guide.html',context=context)

def phd_guide1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(phd_guide.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="phd_guide_report.xlsx"'
    return response



def phd_self1(request,year):
    form = forms.form_phd_self()
    if phd_self.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_phd_self/'+year)
    if request.method == 'POST':

        form = forms.form_phd_self(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'phd_self.html',{'form':form})

def phd_self1_preview(request,year):
    data1 = phd_self.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_phd_self.html',context=context)


def phd_self1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(phd_self.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="phd_self_report.xlsx"'
    return response


def conference_journal1(request,year):
    form = forms.form_conference_journal()
    if conference_journal.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_conference_journal/'+year)
    if request.method == 'POST':

        form = forms.form_conference_journal(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')

    return render(request,'conference_journal.html',{'form':form})


def conference_journal1_preview(request,year):
    data1 = conference_journal.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_conference_journal.html',context=context)


def conference_journal1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(conference_journal.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="conference_journal_report.xlsx"'
    return response



def funded_projects1(request,year):
    form = forms.form_funding()
    if funding.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_funded_projects/'+year)
    if request.method == 'POST':

        form = forms.form_funding(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')
    return render(request,'funded_projects.html',{'form':form})


def funded_projects1_preview(request,year):
    data1 = funding.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_funded_projects.html',context=context)


def funded_projects1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(funding.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="funded_projects_report.xlsx"'
    return response



def open_courses1(request,year):
    form = forms.form_open_courses()
    if open_courses.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_open_courses/'+year)
    if request.method == 'POST':

        form = forms.form_open_courses(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')
    return render(request,'open_courses.html',{'form':form})


def open_courses1_preview(request,year):
    data1 = open_courses.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_open_courses.html',context=context)


def open_courses1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(open_courses.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="open_courses_report.xlsx"'
    return response




def exclusive_research1(request,year):
    form = forms.form_exclusive_research()
    if exclusive_research.objects.filter(info=request.user).filter(year=year):

        return HttpResponseRedirect('/preview_exclusive_research/'+year)
    if request.method == 'POST':

        form = forms.form_exclusive_research(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.info = request.user

            obj.save()

            return HttpResponseRedirect('/index')
    return render(request,'exclusive_research.html',{'form':form})


def exclusive_research1_preview(request,year):
    data1 = exclusive_research.objects.filter(info=request.user).get(year=year)
    context={
            'key':data1
        }
    print(data1)
    return render(request,'preview_exclusive_research.html',context=context)


def exclusive_research1_review(request,dept,year):
    data1 = User.objects.filter(department__name=dept).values()
    dic = {'info':[],
    'name':[]
    }

    for data in data1:
        dic["info"].append(data['id'])

        dic["name"].append(data['first_name'])

    df = pd.DataFrame(list(exclusive_research.objects.filter(info__department__name=dept).filter(year=year).values()))
    df1 = pd.DataFrame.from_dict(dic)

    df2 = pd.merge(df1, df, left_on='info', right_on='info_id')

    df3 = df2.drop(columns=['info','id','info_id'])
    df4 = df3.set_index('name')

    writer = ExcelWriter('PythonExport.xlsx')
    df4.to_excel(writer,'Sheet1')
    writer.save()
    wb = openpyxl.load_workbook('PythonExport.xlsx')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="exclusive_research_report.xlsx"'
    return response
